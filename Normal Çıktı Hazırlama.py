#Doğrulama Kodu
import requests
from bs4 import BeautifulSoup
url = "https://docs.google.com/spreadsheets/d/1AP9EFAOthh5gsHjBCDHoUMhpef4MSxYg6wBN0ndTcnA/edit#gid=0"
response = requests.get(url)
html_content = response.content
soup = BeautifulSoup(html_content, "html.parser")
first_cell = soup.find("td", {"class": "s2"}).text.strip()
if first_cell != "Aktif":
    exit()
first_cell = soup.find("td", {"class": "s1"}).text.strip()
print(first_cell)

import asyncio
import aiohttp
import pandas as pd
import os
import requests
from openpyxl import load_workbook
import time
from colorama import init, Fore, Style
from openpyxl.styles import Font, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy
from openpyxl.worksheet.table import Table, TableStyleInfo
import shutil
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import zipfile
import datetime
from concurrent.futures import ThreadPoolExecutor
from tqdm import tqdm
import gc
pd.options.mode.chained_assignment = None
init(autoreset=True)

print(" ")
print(Fore.BLUE + "https://task.haydigiy.com/admin/exportorder/edit/110/")
print(" ")

parca = int(input(Fore.GREEN + "Kaç Parça İndirilecek: "))
print("E-Tablo Linki: https://docs.google.com/spreadsheets/d/1FJwRFD6ikSsy3uGFRiKp94Iaj1Jd5xerEzJfxJgS1f8/edit#gid=0")
user_input = input("E-tabloda Olan Sipariş Numaraları Hariç Tutulsun mu? (E/H): ").strip().upper()

print(" ")
print("Oturum Açma Başarılı Oldu")
print(" /﹋\ ")
print("(҂`_´)")
print("<,︻╦╤─ ҉ - -")
print("/﹋\\")
print("Mustafa ARI")
print(" ")

#region / Sipariş Listesini İndirme Birleştirme ve Diğer Ayarlar

async def download_file(session, url, index):
    try:
        async with session.get(url, timeout=aiohttp.ClientTimeout(total=99999)) as response:
            if response.status == 200:
                file_name = f"link_{index}.xlsx"
                content = await response.read()
                with open(file_name, "wb") as file:
                    file.write(content)
    except asyncio.TimeoutError:
        print(f"Zaman aşımı hatası: İndirme {index} için zaman aşıma uğradı.")



async def main():
    base_url = "https://www.siparis.haydigiy.com/FaprikaOrderXls/3AG75G/"
    urls = [f"{base_url}{i}/" for i in range(1, parca+1)]

    async with aiohttp.ClientSession() as session:
        tasks = [download_file(session, url, index + 1) for index, url in enumerate(urls)]
        await asyncio.gather(*tasks)

    # Excel dosyalarını birleştirip verileri çıkartma
    data_frames = []
    for i in range(1, parca+1):
        file_name = f"link_{i}.xlsx"
        df = pd.read_excel(file_name)
        data_frames.append(df)

    combined_df = pd.concat(data_frames, ignore_index=True)

    # İstenen sütunları seçme
    selected_columns = ["Id", "OdemeTipi", "TeslimatTelefon", "Barkod", "Adet", "TeslimatEPostaAdresi", "SiparisToplam", "Varyant", "UrunAdi", "KargoTakipNumarasi", "KargoFirmasi"]
    
     # "TeslimatTelefon" sütununda replace yapma
    combined_df['TeslimatTelefon'] = combined_df['TeslimatTelefon'].replace(r'[()/-]', '', regex=True)

    # "TeslimatTelefon" sütunundaki tüm değerleri sayıya çevirme
    combined_df["TeslimatTelefon"] = combined_df["TeslimatTelefon"].apply(pd.to_numeric, errors='coerce')
    final_df = combined_df[selected_columns]
   
    # Virgül ve sonrasını kaldırarak stringi sayıya çevirme işlemi
    def clean_and_convert(value):
        try:
            cleaned_value = value.split(',')[0]  # Virgülden önceki kısmı al
            numeric_value = float(cleaned_value)  # Düzenlenmiş veriyi sayıya çevir
            return numeric_value
        except ValueError:
            return None  # Dönüşüm başarısızsa veya veri boşsa None döndür

    # "Adet" ve "SiparisToplam" sütunlarını temizleme ve dönüştürme
    final_df['Adet'] = final_df['Adet'].apply(clean_and_convert)
    final_df['SiparisToplam'] = final_df['SiparisToplam'].apply(clean_and_convert)
   
    # Düzenlenmiş verileri mevcut dosyanın üzerine kaydetme
    final_df.to_excel("birlesik_excel.xlsx", index=False)
 
    # Birleştirilmiş verileri yeni bir Excel dosyasına kaydetme
    yeni_dosya_adi = "birlesik_excel.xlsx"
    final_df.to_excel(yeni_dosya_adi, index=False)

    # İndirilen dosyaları silme
    for i in range(1, parca+1):
        file_name = f"link_{i}.xlsx"
        if os.path.exists(file_name):
            os.remove(file_name)
           
if __name__ == "__main__":
    asyncio.run(main())

# Kaynak dosya adı
kaynak_excel = "birlesik_excel.xlsx"

# Kopya dosya adı (istediğiniz adı ve konumu belirtin)
kopya_excel = "Kargo Entegrasyonu.xlsx"

# Dosyayı kopyala
shutil.copy(kaynak_excel, kopya_excel)

#endregion

#region / KargoTakipNumarasi ve KargoFirmasi Sütununu Silme

input_file = "birlesik_excel.xlsx"

# Excel dosyasını yükle
df = pd.read_excel(input_file)

# 'KargoTakipNumarasi' sütununu sil
df.drop('KargoTakipNumarasi', axis=1, inplace=True)

# 'KargoTakipNumarasi' sütununu sil
df.drop('KargoFirmasi', axis=1, inplace=True)

# Veri çerçevesini Excel dosyasına kaydet (üzerine yaz)
df.to_excel(input_file, index=False)

#endregion

#region / Hariç Tutulacak Sipariş Numaraları Excelini İndirme ve Hariç Tutma

google_sheet_url = "https://docs.google.com/spreadsheets/d/1FJwRFD6ikSsy3uGFRiKp94Iaj1Jd5xerEzJfxJgS1f8/gviz/tq?tqx=out:csv"

try:
    google_df = pd.read_csv(google_sheet_url)
    google_excel_file = "Hariç Tutulacak Sipariş Numaraları.xlsx"
    google_df.to_excel(google_excel_file, index=False)
except requests.exceptions.RequestException as e:
    pass

def main():
    while True:      
        if user_input == "E":
            # İki Excel dosyasını okuyun
            birlesik_excel = pd.read_excel("birlesik_excel.xlsx")
            haric_excel = pd.read_excel("Hariç Tutulacak Sipariş Numaraları.xlsx")

            # İlk sütunlara göre birleşik Excel verisinden, haric Excel verisinde bulunan satırları filtreleyin
            birlesik_excel = birlesik_excel[~birlesik_excel['Id'].isin(haric_excel['Id'])]

            # Sonucu mevcut "birlesik_excel.xlsx" dosyasına kaydedin (var olan dosyanın üstüne yazacak)
            birlesik_excel.to_excel("birlesik_excel.xlsx", index=False)
            break
        elif user_input == "H":
            break
        else:
            print("Geçerli bir seçenek giriniz (E/H).")

if __name__ == "__main__":
    main()

file_path = "Hariç Tutulacak Sipariş Numaraları.xlsx"

if os.path.exists(file_path):
    os.remove(file_path)

#endregion

#region / Kara Liste Excelini İndirme ve Siparişleri Hariç Tutma

google_sheet_url = "https://docs.google.com/spreadsheets/d/1PgldjEkmmjLPrG9dqvaou481m9QajCOlGxa7wCjwTAQ/gviz/tq?tqx=out:csv"

try:
    google_df = pd.read_csv(google_sheet_url)
    google_excel_file = "Kara Liste.xlsx"
    google_df.to_excel(google_excel_file, index=False)
except requests.exceptions.RequestException as e:
    pass

# "birlesik_excel" dosyasını yükle
birlesik_excel_file = "Kara Liste.xlsx"
sonuc_df = pd.read_excel(birlesik_excel_file)

# "birlesik_excel.xlsx" dosyasını güncelleme
sonuc_df["Telefon Numaraları"] = pd.to_numeric(sonuc_df["Telefon Numaraları"], errors="coerce")  

sonuc_df.to_excel("Kara Liste.xlsx", index=False)







# "birlesik_excel" dosyasını yükle
birlesik_excel_file = "birlesik_excel.xlsx"
birlesik_df = pd.read_excel(birlesik_excel_file)

# "Kara Liste" dosyasını yükle
kara_liste_file = "Kara Liste.xlsx"
kara_liste_df = pd.read_excel(kara_liste_file)

# Çıkarılan satırları tutmak için bir liste oluştur
cikarilan_satirlar = []

# Kara listedeki telefon numaralarını içeren satırları bul, çıkar ve ayrı bir liste'e ekle
for telefon in kara_liste_df.iloc[:, 0]:
    matching_rows = birlesik_df[birlesik_df["TeslimatTelefon"] == telefon]
    
    # ÖdemeTipi sütunu "Kapıda Ödeme" ise işlem yap
    for index, row in matching_rows.iterrows():
        if row["OdemeTipi"] == "Kapıda Ödeme":
            cikarilan_satirlar.append(row)
            birlesik_df.drop(index, inplace=True)

# Çıkarılan satırları içeren bir DataFrame oluştur
cikarilan_satirlar_df = pd.DataFrame(cikarilan_satirlar, columns=birlesik_df.columns)

# Çıkarılan satırları "cikarilan_satirlar.xlsx" dosyasına kaydet
cikarilan_satirlar_df.to_excel("cikarilan_satirlar.xlsx", index=False)

# Güncellenmiş "birlesik_excel" verilerini aynı dosyaya kaydet (üzerine yaz)
birlesik_df.to_excel(birlesik_excel_file, index=False)







# "cikarilan_satirlar" dosyasını yükle
cikarilan_satirlar_file = "cikarilan_satirlar.xlsx"
cikarilan_satirlar_df = pd.read_excel(cikarilan_satirlar_file)

# "Id" ve "TeslimatTelefon" sütunları hariç diğer tüm sütunları sil
sadece_id_telefon_df = cikarilan_satirlar_df[["Id", "TeslimatTelefon"]]
cikarilan_satirlar_df = sadece_id_telefon_df.copy()

# Yenilenen değerleri kaldır (sadece benzersiz satırları bırak)
cikarilan_satirlar_df.drop_duplicates(inplace=True)

try:
    cikarilan_satirlar_df["TeslimatTelefon"] = cikarilan_satirlar_df["TeslimatTelefon"].apply(lambda x: '{:.0f}'.format(x))
except ValueError:
    pass

# Güncellenmiş verileri tekrar kaydet (sadece "Id" ve "TeslimatTelefon" sütunları olacak şekilde)
cikarilan_satirlar_df.to_excel(cikarilan_satirlar_file, index=False)

# "Kara Liste" dosyasını sil
kara_liste_file = "Kara Liste.xlsx"
if os.path.exists(kara_liste_file):
    os.remove(kara_liste_file)


# "cikarilan_satirlar" dosyasının adını "Kara Liste Siparişleri.xlsx" olarak değiştir
yeni_ad = "Kara Liste Siparişleri.xlsx"
os.rename(cikarilan_satirlar_file, yeni_ad)

#endregion

#region / 2500 TL Üzeri Siparişleri Hariç Tutma

# "birlesik_excel" dosyasını yükle
birlesik_excel_file = "birlesik_excel.xlsx"
birlesik_df = pd.read_excel(birlesik_excel_file)

# Belirtilen koşullara uyan satırları filtrele ve sil
condition = ~birlesik_df["TeslimatEPostaAdresi"].str.contains("callcenter.com") & \
            (birlesik_df["OdemeTipi"] == "Kapıda Ödeme") & \
            (birlesik_df["SiparisToplam"] > 2500)

silinecek_satirlar = birlesik_df[condition]
birlesik_df = birlesik_df[~condition]

# Silinen satırları içeren yeni bir DataFrame oluştur
silinecek_df = silinecek_satirlar.copy()

# Güncellenmiş "birlesik_excel" verilerini kaydet
birlesik_df.to_excel(birlesik_excel_file, index=False)

#endregion

#region / Çift Siparişleri Tespit Etme ve Hariç Tutma

# Excel dosyasını oku
input_file = "birlesik_excel.xlsx"
df = pd.read_excel(input_file)

# "OdemeTipi" sütununda "Kapıda Ödeme" olan satırları filtrele
kapida_odeme_df = df[df["OdemeTipi"] == "Kapıda Ödeme"]

# "TeslimatTelefon" ile "Id" sütunlarını birleştirerek "BirlesikVeri" sütununu güncelle
kapida_odeme_df["BirlesikVeri"] = kapida_odeme_df["TeslimatTelefon"].astype(str) + "-" + kapida_odeme_df["Id"].astype(str)

# "BirlesikVeri" sütunundaki değerlerin tekrar sayılarını hesapla ve yeni bir sütun olarak ekle
value_counts = kapida_odeme_df["BirlesikVeri"].value_counts()
kapida_odeme_df["TekrarSayisi"] = kapida_odeme_df["BirlesikVeri"].map(value_counts)

# "TeslimatTelefon" sütunundaki değerlerin tekrar sayılarını hesapla ve yeni bir sütun olarak ekle
value_counts2 = kapida_odeme_df["TeslimatTelefon"].value_counts()
kapida_odeme_df["TekrarSayisi2"] = kapida_odeme_df["TeslimatTelefon"].map(value_counts2)

# "TekrarSayisi" ve "TekrarSayisi2" sütunlarını karşılaştırarak "TEKRAR" sütunu ekleyin
kapida_odeme_df["TEKRAR"] = ["TEKRAR" if ts != ts2 else "" for ts, ts2 in zip(kapida_odeme_df["TekrarSayisi"], kapida_odeme_df["TekrarSayisi2"])]

# Filtreyi kaldır
df = df[df["OdemeTipi"] != "Kapıda Ödeme"]

# Güncellenmiş verileri aynı Excel dosyasına kaydet
df = pd.concat([df, kapida_odeme_df], ignore_index=True)  # Filtrelenmiş verileri ana veri çerçevesine ekleyin
df.to_excel(input_file, index=False, engine="openpyxl")

# Excel dosyasını oku
input_file = "birlesik_excel.xlsx"
df = pd.read_excel(input_file)

# "TEKRAR" yazan satırları seç
tekrar_rows = df[df["TEKRAR"] == "TEKRAR"]

# "TEKRAR" yazan satırları ayrı bir Excel dosyasına kaydet
output_file_tekrar = "Çift Siparişler.xlsx"
tekrar_rows.to_excel(output_file_tekrar, index=False, engine="openpyxl")

# "TEKRAR" yazmayan satırları seç ve "TEKRAR" sütununu kaldır
no_tekrar_rows = df[df["TEKRAR"] != "TEKRAR"]
no_tekrar_rows = no_tekrar_rows.drop(columns=["TEKRAR"])

# Güncellenmiş verileri aynı Excel dosyasına kaydet
no_tekrar_rows.to_excel(input_file, index=False, engine="openpyxl")

# "Çift Siparişler.xlsx" dosyasını oku
tekrar_file = "Çift Siparişler.xlsx"
tekrar_df = pd.read_excel(tekrar_file)

# Sadece "TeslimatTelefon" sütununu tut
tekrar_df = tekrar_df[["TeslimatTelefon"]]

# Temizlenmiş "tekrar_satirlar" verilerini aynı dosyaya kaydet
tekrar_df.to_excel(tekrar_file, index=False, engine="openpyxl")

# Tekrar satırları temizle
tekrar_df.drop_duplicates(subset="TeslimatTelefon", inplace=True)

# "TeslimatTelefon" sütununu sayı biçimine çevir (formatlama)
try:
    tekrar_df["TeslimatTelefon"] = tekrar_df["TeslimatTelefon"].apply(lambda x: '{:.0f}'.format(x))
except ValueError:
    pass

# Temizlenmiş tekrar satırları ayrı bir Excel dosyasına kaydet
output_cleaned_file = "Çift Siparişler.xlsx"
tekrar_df.to_excel(output_cleaned_file, index=False, engine="openpyxl")

#endregion

#region / Hazırlanan Sipariş Numaraları Listesini Verme

# Kaynak dosya adı
kaynak_excel = "birlesik_excel.xlsx"

# Kopya dosya adı (istediğiniz adı ve konumu belirtin)
kopya_excel = "Hazırlanan Sipariş Numaraları.xlsx"

# Dosyayı kopyala
shutil.copy(kaynak_excel, kopya_excel)

#endregion

#region / Birleşik Excel'de Sadece Belirli Sütunları Tutma

# Excel dosyasını okuma
df = pd.read_excel("birlesik_excel.xlsx")

# İşlemi gerçekleştiren fonksiyon
def duplicate_rows(row):
    count = int(row["Adet"])
    return pd.concat([row] * count, axis=1).T

# Tüm satırları işleme tabi tutma
new_rows = df.apply(duplicate_rows, axis=1)

# Yeni veri çerçevesini oluşturma
new_df = pd.concat(new_rows.tolist(), ignore_index=True)

# Sadece belirtilen sütunları seçme
selected_columns = ["Id", "Barkod", "UrunAdi", "Varyant"]
new_df = new_df[selected_columns]

# Veriyi yeni bir Excel dosyasına yazma
new_df.to_excel("birlesik_excel.xlsx", index=False)

#endregion

#region / Haydigiy Online Üzerinden Raf Kodlarını Alma ve Listeye Çektirme

url = "https://haydigiy.online/Products/rafkodlari.php"
response = requests.get(url)
soup = BeautifulSoup(response.text, "html.parser")
table = soup.find("table")
data = []
for row in table.find_all("tr"):
    row_data = []
    for cell in row.find_all(["th", "td"]):
        row_data.append(cell.get_text(strip=True))
    data.append(row_data)
df = pd.DataFrame(data[1:], columns=data[0])
df.to_excel("Raf Kodu.xlsx", index=False)


sonuc_df = pd.read_excel("Raf Kodu.xlsx")

sonuc_df["VaryasyonBarkod"] = pd.to_numeric(sonuc_df["VaryasyonBarkod"], errors="coerce")  # Sayıya dönüştür

# "birlesik_excel.xlsx" dosyasını güncelleme
sonuc_df.to_excel("Raf Kodu.xlsx", index=False)



# "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
sonuc_df = pd.read_excel("birlesik_excel.xlsx")
google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

# "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
sonuc_df["GoogleSheetVerisi"] = "9999-Raf Kodu Yok"

# Her bir "Barkod" değeri için işlem yapma
for index, row in sonuc_df.iterrows():
    barkod = row["Barkod"]
    
    # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama

    matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
    
    # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
    if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 2]):
        sonuc_df.at[index, "GoogleSheetVerisi"] = matching_row.iloc[0, 2]

# "birlesik_excel.xlsx" dosyasını güncelleme
sonuc_df.to_excel("birlesik_excel.xlsx", index=False)

#endregion

#region / Raf Kodlarını Sıralama İçin Düzenleme

# "birlesik_excel.xlsx" dosyasını güncelleme
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi"]  # "GoogleSheetVerisi" sütununu kopyala
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.split("/", n=1).str[0]  # "-" den sonrasını temizle
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.replace(r'^[^-]*-', '', regex=True)
sonuc_df["GoogleSheetVerisi Kopya"] = pd.to_numeric(sonuc_df["GoogleSheetVerisi Kopya"], errors="coerce")  # Sayıya dönüştür
sonuc_df = sonuc_df.sort_values(by="GoogleSheetVerisi Kopya")  # "GoogleSheetVerisi Kopya" sütununa göre sırala
sonuc_df.drop(columns=["GoogleSheetVerisi Kopya"], inplace=True)

sonuc_df.to_excel("birlesik_excel.xlsx", index=False)

# "birlesik_excel.xlsx" dosyasını güncelleme
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi"]  # "GoogleSheetVerisi" sütununu kopyala
sonuc_df["GoogleSheetVerisi Kopya"] = sonuc_df["GoogleSheetVerisi Kopya"].str.split("-", n=1).str[0]  # "-" den sonrasını temizle
sonuc_df["GoogleSheetVerisi Kopya"] = pd.to_numeric(sonuc_df["GoogleSheetVerisi Kopya"], errors="coerce")  # Sayıya dönüştür
sonuc_df = sonuc_df.sort_values(by="GoogleSheetVerisi Kopya")  # "GoogleSheetVerisi Kopya" sütununa göre sırala

sonuc_df.to_excel("birlesik_excel.xlsx", index=False)

#endregion

#region / Raf Kodu Olmayan Kısımları 9999-Raf Kodu Yok İle Doldurma

# "birlesik_excel.xlsx" ve "Raf Kodu.xlsx" dosyalarını okuma
sonuc_df = pd.read_excel("birlesik_excel.xlsx")
google_sheet_df = pd.read_excel("Raf Kodu.xlsx")

# "birlesik_excel.xlsx" dosyasına yeni bir sütun ekleyerek başlangıçta "Raf Kodu Yok" değerleri ile doldurma
sonuc_df["Kategori"] = "9999-Raf Kodu Yok"

# Her bir "Barkod" değeri için işlem yapma
for index, row in sonuc_df.iterrows():
    barkod = row["Barkod"]
    
    # "Raf Kodu.xlsx" dosyasında ilgili "Barkod"u arama
    matching_row = google_sheet_df[google_sheet_df.iloc[:, 0] == barkod]
    
    # Eşleşen "Barkod" varsa ve karşılık gelen hücre boş değilse, değeri "GoogleSheetVerisi" sütununa yazma
    if not matching_row.empty and not pd.isnull(matching_row.iloc[0, 3]):
        sonuc_df.at[index, "Kategori"] = matching_row.iloc[0, 3]

# "birlesik_excel.xlsx" dosyasını güncelleme
sonuc_df.to_excel("birlesik_excel.xlsx", index=False)

file_path = "Raf Kodu.xlsx"

if os.path.exists(file_path):
    os.remove(file_path)

#endregion

#region / İç Giyim Kontrolü

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Yeni Kategori"] = ""

# İç Giyim içeren satırları işle
innerwear_rows = df[df["Kategori"].str.contains("İç Giyim")]

# İç Giyim içeren satırları işle
for index, row in innerwear_rows.iterrows():
    df.loc[index, "Yeni Kategori"] = "İç Giyim"

# Sonucu kaydet
output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
df.to_excel(output_file_path, index=False)

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# "Id" değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")

# "Id" değerlerine göre "Yeni Kategori" değerini güncelle
for group_name, group_data in grouped:
    if any(row["Yeni Kategori"] == "İç Giyim" for _, row in group_data.iterrows()):
        df.loc[df["Id"] == group_name, "Yeni Kategori"] = "İç Giyim"

# Sonucu kaydet
output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
df.to_excel(output_file_path, index=False)









# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# "Yeni Kategori" değeri "İç Giyim" olan satırları seç
innerwear_rows = df[df["Yeni Kategori"] == "İç Giyim"]

# Ayrı Excel dosyasına kaydet
output_file_path = "İç Giyim.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
innerwear_rows.to_excel(output_file_path, index=False)

# Ana DataFrame'den "İç Giyim" satırları sil
df = df[df["Yeni Kategori"] != "İç Giyim"]
df.drop(columns=["Yeni Kategori"], inplace=True)  # "Yeni Kategori" sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / Siparişte Kaç Ürün Olduğunu Tespit Etme

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek tekrar sayılarını tut
df["Tekrar Sayısı"] = df.groupby("Id")["Id"].transform("count")

# Sonucu kaydet
output_file_path = "birlesik_excel.xlsx"  # Sonucun kaydedileceği dosya adı ve yolunu belirtin
df.to_excel(output_file_path, index=False)

#endregion

#region / 0-500 Raf Kodu Arasındaki 14'lü Siparişler (İnstagram 14)

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 0 and max_value <= 500:
        df.loc[df["Id"] == group_name, "Sonuc"] = "0-500-14"

# 0-112 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "0-500-14"]

# Tekrar Sayısı 6 ve daha büyük olan satırları seç
high_repeats_rows = filtered_df[filtered_df.groupby("Id")["Id"].transform("count") >= 5]

# Ayrı Excel dosyasına kaydet
high_repeats_output_file_path = "İnstagram (14).xlsx"
high_repeats_rows.to_excel(high_repeats_output_file_path, index=False)

# Ana DataFrame'den 6 ve üzeri tekrarları çıkar
df.drop(index=high_repeats_rows.index, inplace=True)

# Ana Excel dosyasını güncelle
df.drop(columns=["Sonuc"], inplace=True)
df.to_excel(excel_file_path, index=False)

#endregion

#region / 700-1999 Raf Kodu Arasındaki 14'lü Siparişler (Yeni Depo 14)

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 700 and max_value <= 1999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "700-1999"

# 0-112 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "700-1999"]

# Tekrar Sayısı 6 ve daha büyük olan satırları seç
high_repeats_rows = filtered_df[filtered_df.groupby("Id")["Id"].transform("count") >= 5]

# Ayrı Excel dosyasına kaydet
high_repeats_output_file_path = "Yeni Depo (14).xlsx"
high_repeats_rows.to_excel(high_repeats_output_file_path, index=False)

# Ana DataFrame'den 6 ve üzeri tekrarları çıkar
df.drop(index=high_repeats_rows.index, inplace=True)

# Ana Excel dosyasını güncelle
df.drop(columns=["Sonuc"], inplace=True)
df.to_excel(excel_file_path, index=False)

#endregion

#region / 2000-9999 Raf Kodu Arasındaki 14'lü Siparişler (Özerler Depo 14)

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 2000 and max_value <= 9999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "2000-9999"

# 0-112 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "2000-9999"]

# Tekrar Sayısı 6 ve daha büyük olan satırları seç
high_repeats_rows = filtered_df[filtered_df.groupby("Id")["Id"].transform("count") >= 5]

# Ayrı Excel dosyasına kaydet
high_repeats_output_file_path = "Özerler Depo (14).xlsx"
high_repeats_rows.to_excel(high_repeats_output_file_path, index=False)

# Ana DataFrame'den 6 ve üzeri tekrarları çıkar
df.drop(index=high_repeats_rows.index, inplace=True)

# Ana Excel dosyasını güncelle
df.drop(columns=["Sonuc"], inplace=True)
df.to_excel(excel_file_path, index=False)

#endregion

#region / Geriye Kalan Tüm Depodaki 14'lü Siparişler (Tüm Depo 14)

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 0 and max_value <= 9999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "0-9999"

# 0-112 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "0-9999"]

# Tekrar Sayısı 6 ve daha büyük olan satırları seç
high_repeats_rows = filtered_df[filtered_df.groupby("Id")["Id"].transform("count") >= 5]

# Ayrı Excel dosyasına kaydet
high_repeats_output_file_path = "Tüm Depo (14).xlsx"
high_repeats_rows.to_excel(high_repeats_output_file_path, index=False)

# Ana DataFrame'den 6 ve üzeri tekrarları çıkar
df.drop(index=high_repeats_rows.index, inplace=True)

# Ana Excel dosyasını güncelle
df.drop(columns=["Sonuc"], inplace=True)
df.to_excel(excel_file_path, index=False)

#endregion

#region / 0-112 / 113-206 / 0-206 / 207-400

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# İşlenecek aralıkları belirle
araliklar = [(0, 112), (113, 206), (0, 206), (207, 400)]

# Aralıklar için döngü oluştur
for min_aralik, max_aralik in araliklar:
    # Yeni bir sütun ekleyerek işlem sonuçlarını tut
    df["Sonuc"] = ""

    # Id değerlerine göre grupla ve işlemi yap
    grouped = df.groupby("Id")
    for group_name, group_data in grouped:
        min_value = group_data["GoogleSheetVerisi Kopya"].min()
        max_value = group_data["GoogleSheetVerisi Kopya"].max()

        if min_value >= min_aralik and max_value <= max_aralik:
            df.loc[df["Id"] == group_name, "Sonuc"] = f"{min_aralik}-{max_aralik}"

    # Aralığa uygun olan satırları ayrı bir DataFrame'e kopyala
    filtered_df = df[df["Sonuc"] == f"{min_aralik}-{max_aralik}"]

    # Ayrı Excel dosyasına kaydet
    filtered_output_file_path = f"{min_aralik}-{max_aralik}.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
    filtered_df.to_excel(filtered_output_file_path, index=False)

    # Aralığa uygun olan satırları ana DataFrame'den sil
    df = df[df["Sonuc"] != f"{min_aralik}-{max_aralik}"]
    df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 0-500 Arasındaki Kalan Siparişler (İnstagram Kalanlar) 

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 0 and max_value <= 500:
        df.loc[df["Id"] == group_name, "Sonuc"] = "0-500"

# 0-500 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "0-500"]

# Ayrı Excel dosyasına kaydet
filtered_output_file_path = "İnstagram Kalanlar.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
filtered_df.to_excel(filtered_output_file_path, index=False)

# 0-500 olan satırları ana DataFrame'den sil
df = df[df["Sonuc"] != "0-500"]
df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 700-857 / 858-995 / 996-1133 / 1134-1269 / 700-995 / 996-1269 / 1270-1326 / 1327-1459 / 1460-1531

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# İşlenecek aralıkları belirle
araliklar = [
    (700, 857),
    (858, 995),
    (996, 1133),
    (1134, 1269),
    (700, 995),
    (996, 1269),
    (1270, 1326),
    (1327, 1459),
    (1460, 1531)
]

# Aralıklar için döngü oluştur
for min_aralik, max_aralik in araliklar:
    # Yeni bir sütun ekleyerek işlem sonuçlarını tut
    df["Sonuc"] = ""

    # Id değerlerine göre grupla ve işlemi yap
    grouped = df.groupby("Id")
    for group_name, group_data in grouped:
        min_value = group_data["GoogleSheetVerisi Kopya"].min()
        max_value = group_data["GoogleSheetVerisi Kopya"].max()

        # Aralık kontrolü
        if min_value >= min_aralik and max_value <= max_aralik:
            df.loc[df["Id"] == group_name, "Sonuc"] = f"{min_aralik}-{max_aralik}"

    # Aralığa uygun olan satırları ayrı bir DataFrame'e kopyala
    filtered_df = df[df["Sonuc"] == f"{min_aralik}-{max_aralik}"]

    # Ayrı Excel dosyasına kaydet
    filtered_output_file_path = f"{min_aralik}-{max_aralik}.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
    filtered_df.to_excel(filtered_output_file_path, index=False)

    # Aralığa uygun olan satırları ana DataFrame'den sil
    df = df[df["Sonuc"] != f"{min_aralik}-{max_aralik}"]
    df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 700-1269 Arasındaki Kalan Siparişler (Yeni Depo Kalanlar) 

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 700 and max_value <= 1999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "700-1269"

# 700-1269 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "700-1269"]

# Ayrı Excel dosyasına kaydet
filtered_output_file_path = "Yeni Depo Kalanlar.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
filtered_df.to_excel(filtered_output_file_path, index=False)

# 700-1269 olan satırları ana DataFrame'den sil
df = df[df["Sonuc"] != "700-1269"]
df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 2000-2164 / 2165-2310

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# İşlenecek aralıkları belirle
araliklar = [
    (2000, 2164),
    (2165, 9999)
]

# Aralıklar için döngü oluştur
for min_aralik, max_aralik in araliklar:
    # Yeni bir sütun ekleyerek işlem sonuçlarını tut
    df["Sonuc"] = ""

    # Id değerlerine göre grupla ve işlemi yap
    grouped = df.groupby("Id")
    for group_name, group_data in grouped:
        min_value = group_data["GoogleSheetVerisi Kopya"].min()
        max_value = group_data["GoogleSheetVerisi Kopya"].max()

        # Aralık kontrolü
        if min_value >= min_aralik and max_value <= max_aralik:
            df.loc[df["Id"] == group_name, "Sonuc"] = f"{min_aralik}-{max_aralik}"

    # Aralığa uygun olan satırları ayrı bir DataFrame'e kopyala
    filtered_df = df[df["Sonuc"] == f"{min_aralik}-{max_aralik}"]

    # Ayrı Excel dosyasına kaydet
    filtered_output_file_path = f"{min_aralik}-{max_aralik}.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
    filtered_df.to_excel(filtered_output_file_path, index=False)

    # Aralığa uygun olan satırları ana DataFrame'den sil
    df = df[df["Sonuc"] != f"{min_aralik}-{max_aralik}"]
    df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 2000-2310 Arasındaki Kalan Siparişler (Özerler Depo Kalanlar) 

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 2000 and max_value <= 9999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "2000-9999"

# 700-1269 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "2000-9999"]

# Ayrı Excel dosyasına kaydet
filtered_output_file_path = "Özerler Depo Kalanlar.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
filtered_df.to_excel(filtered_output_file_path, index=False)

# 700-1269 olan satırları ana DataFrame'den sil
df = df[df["Sonuc"] != "2000-9999"]
df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 0-1999 Arasındaki Kalan Siparişler (Yeni Depo ve İnstagram Kalanlar) 

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 0 and max_value <= 1999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "0-1999"

# 700-1269 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "0-1999"]

# Ayrı Excel dosyasına kaydet
filtered_output_file_path = "Yeni Depo ve İnstagram Kalanlar.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
filtered_df.to_excel(filtered_output_file_path, index=False)

# 700-1269 olan satırları ana DataFrame'den sil
df = df[df["Sonuc"] != "0-1999"]
df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / 2000-2310 Arasındaki Kalan Siparişler (İnstagram ve Özerler Depo Kalanlar)

# Excel dosyasını oku
excel_file_path = "birlesik_excel.xlsx"  # Ana Excel dosyasının adını ve yolunu belirtin
df = pd.read_excel(excel_file_path)

# Yeni bir sütun ekleyerek işlem sonuçlarını tut
df["Sonuc"] = ""

# Id değerlerine göre grupla ve işlemi yap
grouped = df.groupby("Id")
for group_name, group_data in grouped:
    min_value = group_data["GoogleSheetVerisi Kopya"].min()
    max_value = group_data["GoogleSheetVerisi Kopya"].max()

    if min_value >= 2000 and max_value <= 9999:
        df.loc[df["Id"] == group_name, "Sonuc"] = "2000-9999"

# 700-1269 olan satırları ayrı bir DataFrame'e kopyala
filtered_df = df[df["Sonuc"] == "2000-9999"]

# Ayrı Excel dosyasına kaydet
filtered_output_file_path = "İnstagram ve Özerler Depo Kalanlar.xlsx"  # Ayrı dosyanın adını ve yolunu belirtin
filtered_df.to_excel(filtered_output_file_path, index=False)

# 700-1269 olan satırları ana DataFrame'den sil
df = df[df["Sonuc"] != "2000-9999"]
df.drop(columns=["Sonuc"], inplace=True)  # Sonuc sütununu sil

# Ana Excel dosyasını güncelle
df.to_excel(excel_file_path, index=False)

#endregion

#region / Geriye Kalan Siparişleri Tüm Depo Kalanlar Yapma

old_file_path = "birlesik_excel.xlsx"
new_file_path = "Tüm Depo Kalanlar.xlsx"

# Dosyanın adını değiştir
os.rename(old_file_path, new_file_path)

#endregion

#region / Raf Kodlarına İkinci Bir Müdahale

# Excel dosyasını aç
file_path = "Tüm Depo Kalanlar.xlsx"
df = pd.read_excel(file_path)
df = pd.read_excel(file_path, engine="openpyxl")

# "GoogleSheetVerisi" sütunundaki "9999-Raf Kodu Yok" değerlerini "Raf Kodu Yok" ile değiştir
df["GoogleSheetVerisi"] = df["GoogleSheetVerisi"].replace("9999-Raf Kodu Yok", "Raf Kodu Yok")

# Değişiklikleri aynı Excel dosyasının üzerine kaydet
df.to_excel(file_path, index=False)

# Excel dosyasını aç
file_path = "İç Giyim.xlsx"
df = pd.read_excel(file_path)
df = pd.read_excel(file_path, engine="openpyxl")

# "GoogleSheetVerisi" sütunundaki "9999-Raf Kodu Yok" değerlerini "Raf Kodu Yok" ile değiştir
df["GoogleSheetVerisi"] = df["GoogleSheetVerisi"].replace("9999-Raf Kodu Yok", "Raf Kodu Yok")

# Değişiklikleri aynı Excel dosyasının üzerine kaydet
df.to_excel(file_path, index=False)

# Excel dosyasını aç
file_path = "Tüm Depo (14).xlsx"
df = pd.read_excel(file_path)
df = pd.read_excel(file_path, engine="openpyxl")

# "GoogleSheetVerisi" sütunundaki "9999-Raf Kodu Yok" değerlerini "Raf Kodu Yok" ile değiştir
df["GoogleSheetVerisi"] = df["GoogleSheetVerisi"].replace("9999-Raf Kodu Yok", "Raf Kodu Yok")

# Değişiklikleri aynı Excel dosyasının üzerine kaydet
df.to_excel(file_path, index=False)

#endregion

#region / Tüm Excellerden Gereksiz Sütunları Temizleme

# İşlenecek Excel dosyalarının listesini oluştur
excel_files = [
    "700-857.xlsx", "700-995.xlsx", "858-995.xlsx", "996-1133.xlsx", "996-1269.xlsx", 
    "1134-1269.xlsx", "Yeni Depo Kalanlar.xlsx", "0-112.xlsx", "113-206.xlsx", "0-206.xlsx", 
    "207-400.xlsx", "1270-1326.xlsx", "1327-1459.xlsx", "1460-1531.xlsx", "2000-2164.xlsx", 
    "2165-2310.xlsx", "Özerler Depo (14).xlsx", "Tüm Depo (14).xlsx", "Yeni Depo (14).xlsx", 
    "İnstagram (14).xlsx", "İç Giyim.xlsx", "İnstagram ve Özerler Depo Kalanlar.xlsx", 
    "Yeni Depo ve İnstagram Kalanlar.xlsx", "Özerler Depo Kalanlar.xlsx", "İnstagram Kalanlar.xlsx", 
    "Tüm Depo Kalanlar.xlsx"
]

# Silinecek sütunlar
columns_to_delete = ["GoogleSheetVerisi Kopya", "Kategori", "Tekrar Sayısı", "Sonuc", "Yeni Kategori"]

# İşleme başla
for file in excel_files:
    try:
        # Excel dosyasını oku
        df = pd.read_excel(file, sheet_name=None)  # Tüm sayfaları okuyalım
        modified_sheets = {}
        
        for sheet_name, data in df.items():
            # Belirtilen sütunlardan olanları sil
            data = data.drop(columns=[col for col in columns_to_delete if col in data.columns], errors='ignore')
            modified_sheets[sheet_name] = data
        
        # Değişiklikleri aynı dosyaya kaydet
        with pd.ExcelWriter(file, engine="openpyxl") as writer:
            for sheet_name, data in modified_sheets.items():
                data.to_excel(writer, sheet_name=sheet_name, index=False)
    
    except Exception:
        # Hataları sessizce geç
        pass

#endregion

#region / 28'li Siparişleri Ayarlama

# İşlenecek Excel dosyalarının listesi
excel_files = [
    "0-112.xlsx", "113-206.xlsx", "0-206.xlsx", "207-400.xlsx",
    "700-857.xlsx", "858-995.xlsx", "996-1133.xlsx", "1134-1269.xlsx",
    "700-995.xlsx", "996-1269.xlsx", "1270-1326.xlsx", "1327-1459.xlsx",
    "1460-1531.xlsx", "2000-2164.xlsx", "2165-9999.xlsx", "İnstagram Kalanlar.xlsx",
    "Yeni Depo Kalanlar.xlsx", "Özerler Depo Kalanlar.xlsx", "Yeni Depo ve İnstagram Kalanlar.xlsx",
    "İnstagram ve Özerler Depo Kalanlar.xlsx", "Tüm Depo Kalanlar.xlsx"
]

for file_name in excel_files:
    # Dosyayı oku
    sonuc_df = pd.read_excel(file_name)

    # "UrunAdi Kopya" sütununu oluştur ve "-" den öncesini temizle
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]

    # "UrunAdiKopya2" sütununu oluştur ve son boşlukla birlikte "-" den sonrasını sil
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(
        lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x
    )

    # "UrunAdiKopya2" sütununu en sağa taşı
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyala ve düzenle
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(
        lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x
    )

    # "UrunAdiKopya3" sütununu en sağa taşı
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # Verileri birleştirip yeni sütun oluştur
    sonuc_df["BirlesikVeri"] = (
        sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]
    )

    # "BirlesikVeri" sütununu en sağa taşı
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "", regex=False)

    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # Değişiklikleri kaydetme
    sonuc_df.to_excel(file_name, index=False)

    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 200
    numbers_per_repeat = 28

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 28
    numbers_per_repeat = 200

    # Verileri ekleme
    for num in range(1, numbers_per_repeat + 1):
        for repeat in range(repeat_count):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    main_sheet = wb["Sheet1"]

    # 1. Sütun başlıklarını değiştirme
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # 2. Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Mevcut verilerle DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları sıralama
    df = df[new_column_order]

    # Başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini tekrar sayfaya yaz
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # 3. Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=1, max_row=main_sheet.max_row, max_col=main_sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 4. Kenarlık ekleme, yazı tipi kalınlaştırma ve boyut ayarlama
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in main_sheet.iter_rows(min_row=1, max_row=main_sheet.max_row, max_col=main_sheet.max_column):
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri tek seferde kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # Diğer sütunları otomatik genişlik olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)  # OpenPyXL sütunları generator döndürür, listeye çeviriyoruz
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # 6. Metni kaydırma formatını hücrelere uygulama
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                wrap_text=True  # Metni kaydır
            )

    # 7. Tablo oluşturma ve stil atama
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    main_sheet.add_table(table)

    main_sheet.column_dimensions["A"].width = 45  # İlk sütun
    main_sheet.column_dimensions["C"].width = 45  # "RAF KODU" sütunu
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydetme
    wb.save(file_name)








# BAT dosyası oluşturma fonksiyonu
def create_bat_files(data, output_folder, batch_size=28):
    batch_count = 1
    remaining_data = data

    while len(remaining_data) > 0:
        current_batch = remaining_data[:batch_size]

        # BAT dosyası oluşturma
        bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
        with open(bat_file_path, "w") as file:
            # İlk link için özel format
            link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
            file.write(link)
            file.write('timeout -t 2\n')  # 2 saniye bekleme süresi

            # Diğer linkleri yazma
            for value in current_batch[1:]:
                link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                file.write(f'start "" {link}\n')

        remaining_data = remaining_data[batch_size:]  # Kalan verileri güncelle
        batch_count += 1

# İşlenen Excel dosyaları üzerinde döngü
for file_name in excel_files:
    # 1. İşlenecek dosyanın adıyla aynı klasör oluştur
    output_folder = os.path.splitext(file_name)[0]  # Dosya adını al (uzantısız)
    os.makedirs(output_folder, exist_ok=True)

    # 2. "Unique Ids" sayfasından "Id" sütunundaki verileri al
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    id_column = unique_ids_sheet["A"][1:]  # İlk satırı atlayarak sadece değerleri al
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # 3. .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # 4. Excel dosyasını ilgili klasöre taşı
    shutil.copy(file_name, os.path.join(output_folder, file_name))

    # 5. Belleği temizle ve orijinal Excel dosyasını sil
    gc.collect()
    os.remove(file_name)

#endregion

#region / 14'lü Siparişleri Ayarlama

# İşlenecek Excel dosyalarının listesi
excel_files = [
    "İç Giyim.xlsx",
    "İnstagram (14).xlsx",
    "Özerler Depo (14).xlsx",
    "Tüm Depo (14).xlsx",
    "Yeni Depo (14).xlsx"
]

for file_name in excel_files:
    # Dosyayı oku
    sonuc_df = pd.read_excel(file_name)

    # "UrunAdi Kopya" sütununu oluştur ve "-" den öncesini temizle
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi"]
    sonuc_df["UrunAdi Kopya"] = sonuc_df["UrunAdi Kopya"].str.split("-", n=1).str[1]

    # "UrunAdiKopya2" sütununu oluştur ve son boşlukla birlikte "-" den sonrasını sil
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdi"]
    sonuc_df["UrunAdiKopya2"] = sonuc_df["UrunAdiKopya2"].apply(
        lambda x: x.split(" - ")[0].rsplit(" ", 1)[-1].strip() if " - " in x else x
    )

    # "UrunAdiKopya2" sütununu en sağa taşı
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya2")
    column_order.append("UrunAdiKopya2")
    sonuc_df = sonuc_df[column_order]

    # "UrunAdi" sütununu tablonun sonuna bir kez daha kopyala ve düzenle
    sonuc_df["UrunAdiKopya3"] = sonuc_df["UrunAdi"].apply(
        lambda x: x.split(" - ", 1)[0].strip() if " - " in x else x
    )

    # "UrunAdiKopya3" sütununu en sağa taşı
    column_order = list(sonuc_df.columns)
    column_order.remove("UrunAdiKopya3")
    column_order.append("UrunAdiKopya3")
    sonuc_df = sonuc_df[column_order]

    # Verileri birleştirip yeni sütun oluştur
    sonuc_df["BirlesikVeri"] = (
        sonuc_df["UrunAdi Kopya"] + " - " + sonuc_df["UrunAdiKopya2"] + " - " + sonuc_df["Varyant"]
    )

    # "BirlesikVeri" sütununu en sağa taşı
    column_order = list(sonuc_df.columns)
    column_order.remove("BirlesikVeri")
    column_order.append("BirlesikVeri")
    sonuc_df = sonuc_df[column_order]

    # "BirlesikVeri" sütunundaki "Beden:" ibaresini çıkarma
    sonuc_df["BirlesikVeri"] = sonuc_df["BirlesikVeri"].str.replace("Beden:", "", regex=False)

    # Belirtilen sütunları silme
    columns_to_drop = ["UrunAdi", "Varyant", "UrunAdi Kopya", "UrunAdiKopya2"]
    sonuc_df.drop(columns_to_drop, axis=1, inplace=True)

    # Değişiklikleri kaydetme
    sonuc_df.to_excel(file_name, index=False)

    # "Id" sütununu teke düşürme
    unique_ids = sonuc_df["Id"].drop_duplicates()

    # Yeni bir Excel sayfası oluşturma ve "Id" değerlerini yazma
    with pd.ExcelWriter(file_name, engine="openpyxl", mode="a") as writer:
        unique_ids.to_excel(writer, sheet_name="Unique Ids", index=False)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 2
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 200
    numbers_per_repeat = 14

    # Verileri ekleme
    for _ in range(repeat_count):
        for num in range(1, numbers_per_repeat + 1):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    sheet = wb["Unique Ids"]

    # Başlangıç sütunu ve satırı
    start_column = 3
    start_row = 2

    # Toplam tekrar sayısı ve her tekrardaki numara adedi
    repeat_count = 14
    numbers_per_repeat = 200

    # Verileri ekleme
    for num in range(1, numbers_per_repeat + 1):
        for repeat in range(repeat_count):
            sheet.cell(row=start_row, column=start_column, value=num)
            start_row += 1

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=1).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # "Id" sütununun verilerini al
    id_column = main_sheet["A"][1:]
    unique_ids_column = unique_ids_sheet["A"][1:]

    # Karşılık gelen değerleri bulup "Sheet1" sayfasının en sağında yeni bir sütuna ekle
    new_column = main_sheet.max_column + 1
    main_sheet.cell(row=1, column=new_column, value="Matching Value (3rd Column)")

    for id_cell in id_column:
        id_value = id_cell.value
        for unique_id_cell in unique_ids_column:
            if unique_id_cell.value == id_value:
                matching_value = unique_id_cell.offset(column=2).value
                main_sheet.cell(row=id_cell.row, column=new_column, value=matching_value)
                break

    # Değişiklikleri kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    main_sheet = wb["Sheet1"]

    # 1. Sütun başlıklarını değiştirme
    new_column_titles = {
        "Id": "SiparişNO",
        "BirlesikVeri": "ÜRÜN",
        "GoogleSheetVerisi": "RAF KODU",
        "UrunAdiKopya3": "ÜRÜN ADI",
        "Matching Value": "KUTU",
        "Matching Value (3rd Column)": "ÇN"
    }

    for col_idx, col_name in enumerate(main_sheet.iter_cols(min_col=1, max_col=main_sheet.max_column), start=1):
        old_title = col_name[0].value
        new_title = new_column_titles.get(old_title, old_title)
        col_name[0].value = new_title

    # 2. Sütunların yeni sıralaması
    new_column_order = [
        "RAF KODU",
        "ÜRÜN",
        "Barkod",
        "KUTU",
        "ÜRÜN ADI",
        "ÇN",
        "SiparişNO"
    ]

    # Mevcut verilerle DataFrame oluştur
    data = main_sheet.iter_rows(min_row=2, values_only=True)
    df = pd.DataFrame(data, columns=[cell.value for cell in main_sheet[1]])

    # Sütunları sıralama
    df = df[new_column_order]

    # Başlıkları güncelle
    for idx, column_name in enumerate(new_column_order, start=1):
        main_sheet.cell(row=1, column=idx, value=column_name)

    # DataFrame verilerini tekrar sayfaya yaz
    for r_idx, row in enumerate(df.values, start=2):
        for c_idx, value in enumerate(row, start=1):
            main_sheet.cell(row=r_idx, column=c_idx, value=value)

    # 3. Hücreleri ortala ve ortaya hizala
    for row in main_sheet.iter_rows(min_row=1, max_row=main_sheet.max_row, max_col=main_sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # 4. Kenarlık ekleme, yazı tipi kalınlaştırma ve boyut ayarlama
    border_style = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    for row in main_sheet.iter_rows(min_row=1, max_row=main_sheet.max_row, max_col=main_sheet.max_column):
        for cell in row:
            cell.font = Font(bold=True, size=14)
            cell.border = border_style

    # Değişiklikleri tek seferde kaydetme
    wb.save(file_name)

    # Sonuç dosyasını yükle
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    main_sheet = wb["Sheet1"]

    # Diğer sütunları otomatik genişlik olarak ayarla
    for column in main_sheet.columns:
        max_length = 0
        column = list(column)  # OpenPyXL sütunları generator döndürür, listeye çeviriyoruz
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        main_sheet.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

    # 6. Metni kaydırma formatını hücrelere uygulama
    for row in main_sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(
                horizontal=cell.alignment.horizontal,
                vertical=cell.alignment.vertical,
                wrap_text=True  # Metni kaydır
            )

    # 7. Tablo oluşturma ve stil atama
    table = Table(displayName="MyTable", ref=main_sheet.dimensions)
    style = TableStyleInfo(
        name="TableStyleLight1",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=True,
    )
    table.tableStyleInfo = style
    main_sheet.add_table(table)

    main_sheet.column_dimensions["A"].width = 45  # İlk sütun
    main_sheet.column_dimensions["C"].width = 45  # "RAF KODU" sütunu
    main_sheet.column_dimensions["G"].width = 14
    main_sheet.column_dimensions["D"].width = 9
    main_sheet.column_dimensions["F"].width = 5

    # Değişiklikleri kaydetme
    wb.save(file_name)

# BAT dosyası oluşturma fonksiyonu
def create_bat_files(data, output_folder, batch_size=14):
    batch_count = 1
    remaining_data = data

    while len(remaining_data) > 0:
        current_batch = remaining_data[:batch_size]

        # BAT dosyası oluşturma
        bat_file_path = os.path.join(output_folder, f"BAT{batch_count}.bat")
        with open(bat_file_path, "w") as file:
            # İlk link için özel format
            link = f'start "" https://task.haydigiy.com/admin/order/printorder/?orderId={current_batch[0]}&isPdf=False\n'
            file.write(link)
            file.write('timeout -t 2\n')  # 2 saniye bekleme süresi

            # Diğer linkleri yazma
            for value in current_batch[1:]:
                link = f"https://task.haydigiy.com/admin/order/printorder/?orderId={value}&isPdf=False"
                file.write(f'start "" {link}\n')

        remaining_data = remaining_data[batch_size:]  # Kalan verileri güncelle
        batch_count += 1

# İşlenen Excel dosyaları üzerinde döngü
for file_name in excel_files:
    # 1. İşlenecek dosyanın adıyla aynı klasör oluştur
    output_folder = os.path.splitext(file_name)[0]  # Dosya adını al (uzantısız)
    os.makedirs(output_folder, exist_ok=True)

    # 2. "Unique Ids" sayfasından "Id" sütunundaki verileri al
    wb = load_workbook(file_name)
    unique_ids_sheet = wb["Unique Ids"]
    id_column = unique_ids_sheet["A"][1:]  # İlk satırı atlayarak sadece değerleri al
    id_values = [cell.value for cell in id_column if cell.value is not None]

    # 3. .bat dosyalarını oluştur ve klasöre kaydet
    create_bat_files(id_values, output_folder)

    # 4. Excel dosyasını ilgili klasöre taşı
    shutil.copy(file_name, os.path.join(output_folder, file_name))

    # 5. Belleği temizle ve orijinal Excel dosyasını sil
    gc.collect()
    os.remove(file_name)


#endregion

#region / Klasörler İçi Boş Kontrolü ve Zip'e Ekleme

#DÜZELTME İÇİN
# Sonuç dosyasını yükle
file_path = "Çift Siparişler.xlsx"
wb = load_workbook(file_path)
main_sheet = wb["Sheet1"]

folders = ["0-112", "0-206", "113-206", "İnstagram (14)", "İnstagram Kalanlar", "İç Giyim", "207-400", "Tüm Depo Kalanlar", "Yeni Depo (14)", "Tüm Depo (14)", "1134-1269", "996-1269", "996-1133", "858-995", "700-995", "700-857", "Yeni Depo Kalanlar", "1270-1326", "1327-1459", "1460-1531", "Özerler Depo (14)", "2000-2164", "2165-9999", "Özerler Depo Kalanlar", "Yeni Depo ve İnstagram Kalanlar", "İnstagram ve Özerler Depo Kalanlar"]

# Bugünkü tarihi al
current_date = datetime.datetime.now().strftime("%Y-%m-%d")

# Oluşturulacak zip dosyasının adı
zip_filename = f"{current_date} Çıktılar.zip"

# Klasörleri kontrol et ve gerektiğinde sil veya zip'e ekle
with zipfile.ZipFile(zip_filename, 'w') as zipf:
    for folder in folders:
        folder_path = os.path.join(".", folder)
        folder_contents = os.listdir(folder_path)
        bat_files = [file for file in folder_contents if file.endswith(".bat")]

        if bat_files:
            for root, _, files in os.walk(folder_path):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.relpath(file_path, "."))
            
            
        else:
            for root, dirs, files in os.walk(folder_path, topdown=False):
                for file in files:
                    file_path = os.path.join(root, file)
                    os.remove(file_path)
                for dir in dirs:
                    dir_path = os.path.join(root, dir)
                    os.rmdir(dir_path)
            os.rmdir(folder_path)

# Klasörleri sil
for folder in folders:
    folder_path = os.path.join(".", folder)
    if os.path.exists(folder_path):
        shutil.rmtree(folder_path)

#endregion

#region / Entegrasyona Gönderme

# Excel dosyasının adını belirtin
excel_dosyasi = "Kargo Entegrasyonu.xlsx"

# Excel dosyasını yükle
df = pd.read_excel(excel_dosyasi)

# "KargoTakipNumarasi" sütununda dolu olan satırları sil
df = df[df['KargoTakipNumarasi'].isna()]


# "KargoFirmasi" sütununda "MNG KARGO" değerine sahip olan satırları sil
df = df[df['KargoFirmasi'] != 'MNG KARGO']

# "KargoFirmasi" sütununda "MNG KARGO" değerine sahip olan satırları sil
df = df[df['KargoFirmasi'] != 'KARGOİST']

# "Id" sütunu hariç diğer tüm sütunları sil
df = df[['Id']]

# Aynı Excel dosyasına kaydet (üzerine yaz)
df.to_excel(excel_dosyasi, index=False)

# Excel dosyasını oku
df = pd.read_excel(excel_dosyasi)

# "Id" sütunu hariç diğer tüm sütunları sil
df = df[['Id']].drop_duplicates()

# Güncellenmiş veriyi aynı Excel dosyasına kaydet (mevcut dosyanın üzerine yazacak)
df.to_excel(excel_dosyasi, index=False)

print(" ")
print(Fore.RED + "Siparişler Entegrasyona Gönderiliyor Ekran Kapanana Kadar İşlem Yapmayın !")

# Excel dosyasını oku
df = pd.read_excel("Kargo Entegrasyonu.xlsx")

# Oturum oluşturma ve giriş yapma işlemini dışarı taşı
session = requests.Session()

# Oturumu bir kez aç
def login():
    login_url = "https://task.haydigiy.com/kullanici-giris/?ReturnUrl=%2Fadmin"
    username = "mustafa_kod@haydigiy.com"
    password = "123456"
    
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36",
        "Referer": "https://task.haydigiy.com/",
    }

    # Oturum açma sayfasına GET isteği
    response = session.get(login_url, headers=headers)
    soup = BeautifulSoup(response.text, "html.parser")

    # __RequestVerificationToken değerini alma
    token = soup.find("input", {"name": "__RequestVerificationToken"}).get("value")

    # Giriş verileri
    login_data = {
        "EmailOrPhone": username,
        "Password": password,
        "__RequestVerificationToken": token,
    }

    # Oturum açma isteği gönderme
    session.post(login_url, data=login_data, headers=headers)

# Oturumu bir kez aç
login()

# Siparişleri gönderme fonksiyonu
def send_request(order_id):
    url = f"https://task.haydigiy.com/admin/order/sendordertoshipmentintegration/?orderId={order_id}"
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/58.0.3029.110 Safari/537.36",
    }

    # Entegrasyon URL'sine istek gönder
    response = session.get(url, headers=headers)



# ThreadPoolExecutor kullanarak istekleri paralel olarak gönderme
with ThreadPoolExecutor(max_workers=5) as executor, tqdm(total=len(df['Id']), desc="Entegrasyona Gönderiliyor") as pbar:
    futures = [executor.submit(send_request, order_id) for order_id in df['Id']]
    for future in futures:
        future.result()  # İşlem tamamlandığında bir sonraki adıma geç
        pbar.update(1)  # İlerleme çubuğunu güncelle


# Excel dosyasının adı
excel_dosyasi = "Kargo Entegrasyonu.xlsx"

# Excel dosyasını sil
try:
    os.remove(excel_dosyasi)
    pass
except FileNotFoundError:
    print(f"{excel_dosyasi} adlı Excel dosyası bulunamadı.")
except Exception as e:
    print(f"Hata oluştu: {e}")

#endregion

#region / Çift Siparişler ve Kara Liste Kontrolü

# Boş dosyaları kontrol etme ve silme işlemi
for dosya in ["Kara Liste Siparişleri.xlsx", "Çift Siparişler.xlsx"]:
    try:
        df = pd.read_excel(dosya) 
        if df.dropna().empty:  
            os.remove(dosya)
        else:
            pass
    except FileNotFoundError:
        print(f"'{dosya}' dosyası bulunamadı.")
    except Exception as e:
        print(f"'{dosya}' dosyasını kontrol ederken bir hata oluştu: {str(e)}")

#endregion

